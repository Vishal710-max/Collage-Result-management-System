import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO
import re

def load_data(path):
    """Load data from session state"""
    return st.session_state.stored_data.get(path, [])

def create_excel_sheet():
    data2 = load_data("Result_dict")
    if not data2:
        st.warning("No detailed data available. Please process a PDF first.")
        return
    
    with st.spinner("Creating Excel sheet..."):
        wb = Workbook()
        ws = wb.active
        ws.title = "Student Results"

        all_codes = []
        for code in data2[0]['Code']:
            if isinstance(code, str):
                all_codes.append(code)

        header = ["Seat No", "Name"]
        for code in all_codes:
            header.extend([code, "UA", "CA", "Total", "Subject_Status"])
        header.extend(["", "Total", "Status", "Percentage"])
        ws.append(header)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        code_indices = {i: code for i, code in enumerate(all_codes)}
        
        for student in data2[:]:
            row = [student["Seat No"], student["Name"]]
            seen_indices = set()
            total_val = 0
            status = ""
            
            if "F" in student.get("Status1", [])[:16]:
                status = "Fail"
            else:
                status = "Pass"

            for val in student["Total"][:9]:
                if val in ('AB', '-', '*'):
                    continue
                elif '*' in val:
                    parts = val.split()
                    if len(parts) == 2 and parts[1].isdigit():
                        total_val += int(parts[1])
                elif '$' in val and '+' in val:
                    match = re.search(r'\$?\s*(\d+)\s*\+\s*(\d+)', val)
                    if match:
                        total_val += int(match.group(1)) + int(match.group(2))
                elif val.isdigit():
                    total_val += int(val)

            percentage = f"{(total_val / 900) * 100:.2f}"

            for code in all_codes:
                if code in code_indices.values():
                    item = [k for k, v in code_indices.items() if v == code and k not in seen_indices]
                    if item: 
                        i = item[0]
                        seen_indices.add(i)
                        row.extend(["", student["UA"][i], student["CA"][i], student["Total"][i], student["Status1"][i]])
                else:
                    row.extend(["", "", "", "", ""])
            
            row.extend(["", total_val, status, percentage])
            ws.append(row)
            current_row = ws.max_row
            ws[f"A{current_row}"].font = Font(bold=True)

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        st.success("Excel sheet created successfully!")
        st.download_button(
            label="📥 Download Excel File",
            data=excel_buffer,
            file_name="BCS-II_Results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

def show():
    st.header("📝 Generate Detailed Excel Report")
    st.info("This will create a comprehensive Excel sheet with all student marks.")
    if st.button("Generate Excel"):
        create_excel_sheet()